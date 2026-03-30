import streamlit as st
import os, requests, json, io, cv2, re
import pandas as pd
import numpy as np
from PIL import Image
from dotenv import load_dotenv
from groq import Groq
from datetime import datetime, timedelta
from streamlit_cookies_manager import EncryptedCookieManager

# --- CONFIGURATION & AUTH ---
st.set_page_config(page_title="Inventory OS", page_icon="📦", layout="centered")

# Native Cloud Secrets: Safe for Deployment
GROQ_API_KEY = st.secrets.get("GROQ_API_KEY") or os.getenv("GROQ_API_KEY")
OCR_SPACE_API_KEY = st.secrets.get("OCR_SPACE_API_KEY") or os.getenv("OCR_SPACE_API_KEY", "helloworld")
EXCEL_FILE = "inventory.xlsx"
TXN_FILE = "transactions.csv"
THRESHOLD_FILE = "thresholds.json"

def load_thresholds():
    if os.path.exists(THRESHOLD_FILE):
        with open(THRESHOLD_FILE, "r") as f:
            return json.load(f)
    return {}

def save_thresholds(t_dict):
    with open(THRESHOLD_FILE, "w") as f:
        json.dump(t_dict, f, indent=2)

# 1. Persistent Login System (6-hour Encrypted Cookies)
SESSION_HOURS = 6
cookies = EncryptedCookieManager(
    prefix="inv_os_",
    password=st.secrets.get("COOKIE_SECRET") or os.getenv("COOKIE_SECRET", "godrej-inv-secret-key-2024")
)
if not cookies.ready():
    st.stop()  # Wait for cookies to load (synchronous & reliable)

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

    # Try Cookie Recovery
    c_user = cookies.get("inv_user")
    c_time = cookies.get("inv_login_time")
    if c_user and c_time:
        try:
            login_dt = datetime.fromisoformat(c_time)
            if datetime.now() - login_dt < timedelta(hours=SESSION_HOURS):
                with open("users.json", "r") as f:
                    v_users = json.load(f)
                if c_user in v_users:
                    st.session_state["authenticated"] = True
                    st.session_state["user"] = c_user
                    st.session_state["name"] = v_users[c_user]["name"]
                    st.session_state["page"] = "Inventory" if c_user == "admin" else "Update Stock"
        except Exception:
            pass  # Corrupt cookie, just show login

if not st.session_state["authenticated"]:
    st.title("🔒 Staff Login")
    with st.form("Login"):
        username = st.text_input("Username").strip()
        password = st.text_input("Password", type="password")
        if st.form_submit_button("Login"):
            with open("users.json", "r") as f:
                v_users = json.load(f)
            if username in v_users and v_users[username]["password"] == password:
                # Set persistent cookies (6-hour session)
                cookies["inv_user"] = username
                cookies["inv_login_time"] = datetime.now().isoformat()
                cookies.save()
                st.session_state["authenticated"] = True
                st.session_state["user"] = username
                st.session_state["name"] = v_users[username]["name"]
                st.session_state["page"] = "Inventory" if username == "admin" else "Update Stock"
                st.session_state["transaction_type"] = None
                st.rerun()
            else:
                st.error("Invalid Credentials!")
    st.stop()

# --- CACHED RESOURCES ---
PRICE_CACHE = "_price_list_cache.pkl"

@st.cache_data
def load_price_list():
    """Smart loader that handles varied column names (LN Code vs Product Code)."""
    try:
        # 1. Excel File Check
        xls_path = "priceListHomeFurniture.xlsx"
        if not os.path.exists(xls_path): return None
        
        # 2. Cache Check
        if os.path.exists(PRICE_CACHE):
            if os.path.getmtime(PRICE_CACHE) > os.path.getmtime(xls_path):
                return pd.read_pickle(PRICE_CACHE)
        
        # 3. Flexible Column Synonyms
        CODE_ALIAES = ["LN Code", "Product Code", "Item Code", "Code"]
        DESC_ALIAES = ["LN Description", "Product Name", "Description", "Item Name"]
        
        xls = pd.ExcelFile(xls_path)
        all_dfs = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=5)
            
            # Find Best Columns
            col_code = next((c for c in CODE_ALIAES if c in df.columns), None)
            col_desc = next((c for c in DESC_ALIAES if c in df.columns), None)
            
            if col_code and col_desc:
                subset = df[[col_code, col_desc]].copy()
                # Use standardized internal names
                subset.columns = ["LN Code", "LN Description"]
                subset["Unit Consumer Basic"] = df.get("Unit Consumer Basic", pd.Series(dtype="object"))
                subset["Category"] = sheet_name
                all_dfs.append(subset)
        
        if all_dfs:
            combined = pd.concat(all_dfs, ignore_index=True)
            combined = combined.dropna(subset=["LN Code"])
            combined["LN Code"] = combined["LN Code"].astype(str).str.strip()
            combined.to_pickle(PRICE_CACHE)
            return combined
        return None
    except Exception as e:
        st.session_state["last_ocr_err"] = f"Price List Loading Error: {e}"
        return None
        st.sidebar.error(f"Price List Missing or Error: {e}")
        return None

price_list_df = load_price_list()

# --- SIDEBAR UI ---
st.sidebar.title(f"👤 Welcome, {st.session_state.get('name', 'Staff')}")
st.sidebar.divider()

# Navigation
if st.sidebar.button("📦 View Inventory", use_container_width=True):
    st.session_state["page"] = "Inventory"
    st.rerun()

if st.sidebar.button("🔄 Update Stock", use_container_width=True):
    st.session_state["page"] = "Update Stock"
    st.session_state["transaction_type"] = None
    st.rerun()

if st.sidebar.button("📜 View History", use_container_width=True):
    st.session_state["page"] = "History"
    st.rerun()

if st.session_state.get("user") == "admin":
    if st.sidebar.button("🛠️ Manage Staff", use_container_width=True):
        st.session_state["page"] = "Manage Staff"
        st.rerun()
    if st.sidebar.button("📉 Stock Maintenance", use_container_width=True):
        st.session_state["page"] = "Stock Maintenance"
        st.rerun()
    if st.sidebar.button("🛠️ Manage Master List", use_container_width=True):
        st.session_state["page"] = "Manage Master List"
        st.rerun()

st.sidebar.divider()
if st.session_state.get("user") == "admin":
    if st.sidebar.button("🗑️ Clear Inventory", type="secondary"):
        for f in [EXCEL_FILE, TXN_FILE, PRICE_CACHE]:
            if os.path.exists(f):
                os.remove(f)
        st.cache_data.clear()
        st.sidebar.success("Inventory & transactions cleared!")
        st.rerun()

if st.sidebar.button("Logout", use_container_width=True):
    # Safely clear persistent cookies
    try:
        cookies["inv_user"] = ""
        cookies["inv_login_time"] = ""
        cookies.save()
    except Exception:
        pass
    st.session_state["authenticated"] = False
    st.rerun()

if st.session_state.get("user") == "admin":
    with st.sidebar.expander("🛡️ Data Controls"):
        if st.button("🧹 Cleanup Rogue Data", help="Scans for 1,000,000+ unit errors", use_container_width=True):
            if os.path.exists(TXN_FILE):
                try:
                    cdf = pd.read_csv(TXN_FILE)
                    # Check for huge Qty Diff
                    bad_mask = cdf["Qty Diff"].abs() > 5000 
                    bad_count = bad_mask.sum()
                    if bad_count > 0:
                        cdf = cdf[~bad_mask]
                        cdf.to_csv(TXN_FILE, index=False)
                        st.sidebar.success(f"Fixed {bad_count} massive entries!")
                        st.rerun()
                    else:
                        st.sidebar.info("All good! No massive entries found.")
                except Exception as e:
                    st.sidebar.error(f"Error: {e}")
            else:
                st.sidebar.info("No transaction file found.")

if st.session_state.get("user") == "admin":
    with st.sidebar.expander("🛠️ OCR Diagnostic"):
        st.write("**Extracted Text:**")
        st.code(st.session_state.get("raw_ocr_code", "No extraction yet"))
        st.write("**Status:**")
        st.code(st.session_state.get("last_ocr_err", "No errors recorded"))
        if st.button("Reset Log"):
            st.session_state["raw_ocr_code"] = "None"
            st.session_state["last_ocr_err"] = "No errors recorded"
            st.rerun()

# --- CORE FUNCTIONS ---
def sanitize_quantity(val):
    """Safety guard: no single auto-extraction should exceed 500 units."""
    try:
        qty = int(val)
        if qty > 500: return 1 # Reset to 1 if it smells like a phone number
        return max(1, qty)
    except Exception:
        return 1

def hard_extract_math(text):
    """Primary regex for the 15-character 8+SD+5 format."""
    pc, qty = None, None
    # 1. Product Code: 8 digits, SD, 5 digits
    pc_match = re.search(r"(\d{7,9})[A-Z0-9]{1,2}(\d{4,6})", text, re.IGNORECASE)
    if pc_match:
        raw_pre, raw_suf = pc_match.group(1), pc_match.group(2)
        pc = f"{raw_pre[-8:]}SD{raw_suf[:5]}"
        
    # 2. QTY: Using \b (boundaries) to avoid matching phone numbers
    qty_match = re.search(r"(?:NET\s*QUANTITY|UNITS?|QTY)[\s:]*\b(\d+)\b", text, re.IGNORECASE)
    if qty_match: qty = sanitize_quantity(qty_match.group(1))
    return pc, qty

def repair_ocr_code(text):
    """Fuzzy repair for the 15-digit pattern if hard_extract_math missed it."""
    # Look for any sequence of 13-17 chars that smells like a code
    potential = re.findall(r"[A-Z0-9]{13,17}", text.upper())
    for p in potential:
        # If it has an 'S' or 'D' in the middle, try to force it to 8+SD+5
        m = re.search(r"(\d{7,9})[SD]{1,2}(\d{4,6})", p)
        if m:
            return f"{m.group(1)[-8:]}SD{m.group(2)[:5]}"
    return None

def sharpen_image(pil_img):
    img = np.array(pil_img)
    # Balanced sharpen: improves 8/3/1/4 accuracy without creating blooming
    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
    return Image.fromarray(cv2.filter2D(img, -1, kernel))

def scan_document(pil_image):
    image = np.array(pil_image)
    if len(image.shape) == 3 and image.shape[2] == 4:
        image = cv2.cvtColor(image, cv2.COLOR_RGBA2RGB)
    
    orig = image.copy()
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    
    grad = cv2.morphologyEx(gray, cv2.MORPH_GRADIENT, cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)))
    _, thresh = cv2.threshold(grad, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_RECT, (50, 15)))
    closed = cv2.morphologyEx(closed, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_RECT, (15, 50)))
    
    cnts, _ = cv2.findContours(closed.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts: return pil_image
    
    # Text-Density Geometry: Find the densest text-block area
    largest = max(cnts, key=cv2.contourArea)
    if cv2.contourArea(largest) < (image.shape[0] * image.shape[1] * 0.015): # 1.5% minimum area
        return pil_image
        
    x, y, w, h = cv2.boundingRect(largest)
    pad_y = int(h * 0.12)
    pad_x = int(w * 0.12)
    
    y1 = max(0, y - pad_y)
    y2 = min(image.shape[0], y + h + pad_y)
    x1 = max(0, x - pad_x)
    x2 = min(image.shape[1], x + w + pad_x)
    
    return Image.fromarray(orig[y1:y2, x1:x2])

def compress_image(pil_img, max_size=1280):
    """Resize image to a manageable size for mobile/OCR performance."""
    if max(pil_img.size) > max_size:
        ratio = max_size / float(max(pil_img.size))
        new_size = (int(pil_img.size[0] * ratio), int(pil_img.size[1] * ratio))
        return pil_img.resize(new_size, Image.LANCZOS)
    return pil_img

# --- CORE UTILS ---
client = Groq(api_key=GROQ_API_KEY)

def get_ocr_text(img_t):
    """Run OCR on a PIL image and return cleaned text."""
    img_byte_arr = io.BytesIO()
    img_t.save(img_byte_arr, format='JPEG', quality=85)
    files = {"file": ("image.jpg", img_byte_arr.getvalue(), "image/jpeg")}
    data = {"apikey": OCR_SPACE_API_KEY, "language": "eng", "isTable": True, "scale": True, "OCREngine": 2, "isOverlayRequired": False, "filetype": "JPG", "detectOrientation": True}
    try:
        response = requests.post("https://api.ocr.space/parse/image", files=files, data=data, timeout=20)
        if response.status_code == 200:
            ocr_result = response.json()
            if not ocr_result.get("IsErroredOnProcessing"):
                pr = ocr_result.get("ParsedResults", [])
                if pr:
                    t = pr[0].get("ParsedText", "").strip()
                    if t:
                        return re.sub(r'(?:MRP|USP)\s*[\:R]*\s*7\s+(\d+)', r'MRP \1', t, flags=re.IGNORECASE)
                else:
                    st.session_state["last_ocr_err"] = "No Parsed Results in JSON"
            else:
                st.session_state["last_ocr_err"] = ocr_result.get("ErrorMessage", "Unknown OCR processing error")
        else:
            st.session_state["last_ocr_err"] = f"OCR.space API Error {response.status_code}"
    except Exception as e:
        st.session_state["last_ocr_err"] = str(e)
    return ""

def extract_from_image(uploaded_file):
    """Multi-Stage Extraction with Full-Image Fallback."""
    try:
        uploaded_file.seek(0)
        orig_pil = Image.open(uploaded_file)
        
        # --- STAGE 1: THE CROP ---
        proc_img_crop = sharpen_image(scan_document(compress_image(orig_pil)))
        text_crop = get_ocr_text(proc_img_crop)
        res_crop = parse_and_lookup(text_crop) if text_crop else None
        
        # If Stage 1 found a valid Product In Price List, Return it
        if res_crop and res_crop.get("product_name") != "Unknown Product":
            st.session_state["last_ocr_err"] = "Scan Success (Stage: Crop)"
            return res_crop
            
        # --- STAGE 2: THE FALLBACK (Full Image) ---
        st.session_state["last_ocr_err"] = "Fallback: Trying Full Image (Product Not in List)"
        proc_img_full = sharpen_image(compress_image(orig_pil))
        text_full = get_ocr_text(proc_img_full)
        res_full = parse_and_lookup(text_full) if text_full else None
        
        if res_full:
            st.session_state["last_ocr_err"] = "Scan Success (Stage: Full Image)"
            return res_full
            
    except Exception as e:
        st.session_state["last_ocr_err"] = f"Extraction System Error: {e}"
    
    return None

def parse_and_lookup(text):
    """Parse text and verify against Golden Database using Similarity."""
    # 1. Pattern-Based Extraction
    pc, qty = hard_extract_math(text)
    if not pc:
        pc = repair_ocr_code(text)
    
    # 2. LLM Extraction (Strict JSON Only)
    prompt = f"JSON ONLY. Fields: product_code, quantity (int), mrp (num), package. (NO DIMENSIONS). Text: {text}"
    try:
        res = client.chat.completions.create(
            messages=[{"role": "system", "content": "Return ONLY JSON with 4 keys: product_code, quantity, mrp, package. Forbidden: dimensions, width, height, length, material."}, {"role": "user", "content": prompt}],
            model="llama-3.1-8b-instant", temperature=0.1, response_format={"type": "json_object"}
        )
        ext_data = json.loads(res.choices[0].message.content)
        # Explicit clean: throw away any rogue fields
        allowed_keys = ["product_code", "quantity", "mrp", "package"]
        ext_data = {k: ext_data[k] for k in allowed_keys if k in ext_data}
    except Exception:
        ext_data = {}

    final_pc = pc if pc else ext_data.get("product_code")
    if not final_pc: return None
    
    final_pc = str(final_pc).upper().replace(" ", "").strip()
    st.session_state["raw_ocr_code"] = final_pc # Log for diagnostics

    # 3. Golden Database Lookup & Fuzzy Similarity
    p_name = "Unknown Product"
    base_price = "N/A"
    p_category = "N/A"

    if price_list_df is not None:
        p_code_str = str(final_pc)
        # Try Exact
        matches = price_list_df[price_list_df["LN Code"] == p_code_str]
        
        # Try Fuzzy Correction (Swaps)
        if matches.empty:
            swaps = {"3": "8", "8": "3", "S": "5", "5": "S", "0": "O", "O": "0", "1": "I", "I": "1"}
            for i, char in enumerate(p_code_str):
                if char in swaps:
                    test_code = p_code_str[:i] + swaps[char] + p_code_str[i+1:]
                    matches = price_list_df[price_list_df["LN Code"] == test_code]
                    if not matches.empty:
                        final_pc = test_code
                        break

        # Try Similarity (Levenshtein) - If still not found
        if matches.empty:
            from difflib import SequenceMatcher
            best_sc = 0
            best_code = None
            # Only search nearby if the code is at least 10 chars
            if len(p_code_str) > 10:
                for db_code in price_list_df["LN Code"].head(5000): # Limit search to avoid lag
                    ratio = SequenceMatcher(None, p_code_str, db_code).ratio()
                    if ratio > 0.85 and ratio > best_sc:
                        best_sc = ratio
                        best_code = db_code
                
                if best_code:
                    final_pc = best_code
                    matches = price_list_df[price_list_df["LN Code"] == final_pc]

        if not matches.empty:
            row = matches.iloc[0]
            p_name = str(row.get("LN Description", p_name))
            base_price = str(row.get("Unit Consumer Basic", base_price))
            p_category = str(row.get("Category", p_category))

    qty_val = qty if qty is not None else sanitize_quantity(ext_data.get("quantity", 1))
    
    return {
        "product_code": str(final_pc), "product_name": p_name, "quantity": qty_val,
        "mrp": ext_data.get("mrp"), "package": ext_data.get("package", "N/A"), "base_price": base_price, "category": p_category
    }

# --- PAGES ---
current_page = st.session_state.get("page", "Inventory")

# 1. INVENTORY PAGE
if current_page == "Inventory":
    st.title("📦 Premium Inventory Dashboard")
    
    if os.path.exists(EXCEL_FILE):
        try:
            # Aggregate all sheets into one
            all_inv_dfs = []
            xls_display = pd.ExcelFile(EXCEL_FILE)
            for sn in xls_display.sheet_names:
                if sn.lower() not in ("sheet1", "sheet 1"):
                    sdf = pd.read_excel(xls_display, sheet_name=sn)
                    sdf["Category"] = sn
                    all_inv_dfs.append(sdf)
            
            if not all_inv_dfs:
                st.info("Inventory is empty.")
                st.stop()
                
            main_df = pd.concat(all_inv_dfs, ignore_index=True)
            thresholds = load_thresholds()
            
            # --- METRICS ---
            col1, col2, col3 = st.columns(3)
            total_sku = len(main_df["Product Code"].unique())
            total_qty = int(main_df["Quantity"].sum())
            
            # Count Low Stock items using thresholds
            low_stock_count = 0
            for _, row in main_df.iterrows():
                pc = str(row["Product Code"])
                t_val = thresholds.get(pc, 0)
                thresh = t_val.get("min", 0) if isinstance(t_val, dict) else t_val
                if thresh > 0 and row["Quantity"] < thresh: low_stock_count += 1
                
            col1.metric("Total items", total_sku)
            col2.metric("Total units", total_qty)
            col3.metric("Low Stock Alerts", low_stock_count, delta=-low_stock_count, delta_color="inverse")

            st.divider()
            
            # --- SEARCH & FILTER ---
            search_query = st.text_input("🔍 Search Inventory", placeholder="Search by Code or Product Name...").lower()
            cat_filter = st.multiselect("🏷️ Filter by Category", main_df["Category"].unique())
            
            # Apply Filters
            filtered_df = main_df.copy()
            if search_query:
                filtered_df = filtered_df[
                    filtered_df["Product Code"].astype(str).str.lower().str.contains(search_query) |
                    filtered_df["Product Name"].astype(str).str.lower().str.contains(search_query)
                ]
            if cat_filter:
                filtered_df = filtered_df[filtered_df["Category"].isin(cat_filter)]
            
            # --- BETTER READABILITY STYLING ---
            def stock_heatmap(row):
                pc = str(row["Product Code"])
                t_val = thresholds.get(pc, 0)
                thresh = t_val.get("min", 0) if isinstance(t_val, dict) else t_val
                qty = row["Quantity"]
                
                # Base styles (Normal)
                styles = [""] * len(row)
                qty_idx = list(filtered_df.columns).index("Quantity")
                
                if qty == 0:
                    # Clear Red warning for Out of Stock (Whole Row)
                    return ["background-color: #fee2e2; color: #991b1b"] * len(row)
                elif qty < thresh:
                    # Subtle Orange warning for Low Stock (Only the Quantity cell)
                    styles[qty_idx] = "background-color: #fef3c7; color: #92400e; font-weight: bold"
                
                return styles

            if not filtered_df.empty:
                styled_df = filtered_df.style.apply(stock_heatmap, axis=1)
                st.dataframe(styled_df, use_container_width=True, hide_index=True)
            else:
                st.info("No matching items found.")

        except Exception as e:
            st.error(f"Error loading inventory dashboard: {e}")
    else:
        st.info("Inventory is empty. Start by updating stock!")

# 2. HISTORY PAGE
elif current_page == "History":
    st.title("📜 Transaction History")
    
    # 1. Date Selector
    sel_date = st.date_input("📅 Select Date", datetime.now().date())
    
    if os.path.exists(TXN_FILE):
        try:
            txn_df = pd.read_csv(TXN_FILE)
            if txn_df.empty:
                st.info("No history recorded yet.")
                st.stop()

            # Ensure minimal columns exist
            for col in ["Timestamp", "User", "Status"]:
                if col not in txn_df.columns: txn_df[col] = "N/A"

            # Filter by Date
            txn_df['Date'] = pd.to_datetime(txn_df['Timestamp']).dt.date
            day_df = txn_df[txn_df['Date'] == sel_date].copy()

            if st.session_state["user"] != "admin":
                # Staff see only their own for that day
                display_df = day_df[day_df["User"].astype(str).str.strip() == str(st.session_state["user"]).strip()].copy()
                st.info(f"Showing history for {st.session_state['name']} on {sel_date}")
            else:
                # Admin see all for that day
                display_df = day_df.copy()
                st.info(f"Administrator View: Showing all transactions for {sel_date}")
                
                # Admin Deletion Button for the day
                if not display_df.empty:
                    if st.button(f"🗑️ Clear All History for {sel_date}", type="secondary", use_container_width=True):
                        # Filter OUT the rows for this date in the ORIGINAL txn_df
                        remaining_df = txn_df[txn_df['Date'] != sel_date]
                        remaining_df.drop(columns=['Date'], inplace=True)
                        remaining_df.to_csv(TXN_FILE, index=False)
                        st.success(f"History for {sel_date} cleared!")
                        st.rerun()
            
            if display_df.empty:
                st.write(f"No transactions found for {sel_date}.")
            else:
                # Column sanitizing for sort
                sort_col = "Timestamp" if "Timestamp" in display_df.columns else display_df.columns[0]
                display_df = display_df.sort_values(sort_col, ascending=False)
                if 'Date' in display_df.columns: display_df.drop(columns=['Date'], inplace=True)

                # Style failed/partial entries
                def style_status(row):
                    color = ""
                    if str(row.get("Status", "")).lower() == "failed": color = "background-color: #ffcccc"
                    elif str(row.get("Status", "")).lower() == "partial": color = "background-color: #fff4cc"
                    return [color] * len(row)

                styled_df = display_df.style.apply(style_status, axis=1)
                st.dataframe(styled_df, use_container_width=True, hide_index=True)
        except Exception as e:
            st.error(f"Error reading history: {e}")
    else:
        st.info("No history recorded yet.")

# 3. STOCK MAINTENANCE PAGE (Admin Only)
elif current_page == "Stock Maintenance" and st.session_state.get("user") == "admin":
    st.title("📉 Minimum Stock Maintenance")
    thresholds = load_thresholds()
    
    # 1. Add/Update Threshold Form
    st.subheader("⚙️ Set Product Threshold")
    if price_list_df is not None:
        # Create search labels: "Code | Description"
        p_options = [f"{row['LN Code']} | {row['LN Description']}" for _, row in price_list_df.iterrows()]
        sel_raw = st.selectbox("Search Product from Price List", [""] + p_options, help="Type to filter products")
        
        if sel_raw:
            p_code = sel_raw.split(" | ")[0]
            p_name = sel_raw.split(" | ")[1]
            t_val = thresholds.get(p_code, 0)
            curr_t = t_val.get("min", 0) if isinstance(t_val, dict) else t_val
            
            with st.form("Set Threshold"):
                st.write(f"Editing: **{p_name}** (`{p_code}`)")
                new_t = st.number_input("Desired Minimum Stock Level", min_value=0, value=curr_t)
                if st.form_submit_button("Confirm & Save"):
                    thresholds[p_code] = {"name": p_name, "min": new_t}
                    save_thresholds(thresholds)
                    st.success(f"Threshold for {p_code} set to {new_t}!")
                    st.rerun()

    # 2. View All Thresholds
    st.divider()
    st.subheader("📋 Minimum Stock Values")
    
    if thresholds:
        # Loop through existing thresholds to add Edit/Delete rows
        for k, v in list(thresholds.items()):
            # Canonicalize data (handle old int or new dict)
            name = v.get("name", "N/A") if isinstance(v, dict) else "N/A"
            min_val = v.get("min", 0) if isinstance(v, dict) else v
            
            row_col1, row_col2 = st.columns([3, 1])
            with row_col1:
                st.write(f"**{name}** (`{k}`): **{min_val}**")
            with row_col2:
                with st.popover("⋮"):
                    # Edit Option
                    st.write(f"Edit Threshold: **{name}**")
                    new_edit_t = st.number_input("New Min Level", min_value=0, value=min_val, key=f"edit_val_{k}")
                    if st.button("Update Value", key=f"btn_edit_{k}", use_container_width=True, type="primary"):
                        thresholds[k] = {"name": name, "min": new_edit_t}
                        save_thresholds(thresholds)
                        st.toast(f"Updated {k} to {new_edit_t}")
                        st.rerun()
                    
                    st.divider()
                    # Delete Option
                    if st.button("🗑️ Delete", key=f"btn_del_{k}", use_container_width=True, type="secondary"):
                        del thresholds[k]
                        save_thresholds(thresholds)
                        st.toast(f"Threshold for {k} deleted.")
                        st.rerun()
    else:
        st.info("No custom thresholds set. Alerting on 0 quantity only.")

# 4. MANAGE STAFF PAGE (Admin Only)
elif current_page == "Manage Staff" and st.session_state.get("user") == "admin":
    st.title("🛠️ Manage Staff Accounts")
    
    # Reload users.json to ensure freshness
    with open("users.json", "r") as f:
        current_users = json.load(f)
    
    # 1. List Current Staff
    st.subheader("📋 Current Accounts")
    
    for uid, info in list(current_users.items()):
        if uid == "admin": continue
        col1, col2 = st.columns([3, 1])
        with col1:
            st.write(f"**{info['name']}** (`{uid}`)")
        with col2:
            with st.popover("⋮"):
                if st.button("🗑️ Delete Account", key=f"del_{uid}", use_container_width=True, type="secondary"):
                    del current_users[uid]
                    with open("users.json", "w") as f:
                        json.dump(current_users, f, indent=2)
                    st.toast(f"Account for '{uid}' deleted.")
                    st.rerun()

    st.divider()
    
    # 2. Add New Staff member
    st.subheader("👤 Register New Staff")
    with st.form("Add User"):
        new_uid = st.text_input("Username (e.g. staff2)").strip().lower()
        new_name = st.text_input("Full Name (e.g. John Doe)").strip()
        new_pass = st.text_input("Password", type="password").strip()
        
        if st.form_submit_button("Register Account"):
            if not new_uid or not new_name or not new_pass:
                st.error("All fields are required!")
            elif new_uid in current_users:
                st.error(f"Username '{new_uid}' already exists!")
            else:
                current_users[new_uid] = {"password": new_pass, "name": new_name}
                with open("users.json", "w") as f:
                    json.dump(current_users, f, indent=2)
                st.success(f"Success! Account for {new_name} is now active.")
                st.rerun()

# 4. UPDATE STOCK PAGE
elif current_page == "Update Stock":
    st.title("🔄 Update Stock")
    
    # Step 1: Big Buttons for Selection
    if st.session_state.get("transaction_type") is None:
        col1, col2 = st.columns(2)
        if col1.button("➕ INCOMING\n(Add Stock)", type="primary", use_container_width=True):
            st.session_state["transaction_type"] = "Incoming"
            st.rerun()
        if col2.button("➖ OUTGOING\n(Reduce Stock)", type="primary", use_container_width=True):
            st.session_state["transaction_type"] = "Outgoing"
            st.rerun()
    else:
        transaction_type = st.session_state["transaction_type"]
        st.info(f"Mode: **{transaction_type.upper()}** (Tap 'Update Stock' in menu to change)")
        
        uploaded_files = st.file_uploader("Drop Scan/Label Images Here", type=["jpg", "png", "jpeg", "webp"], accept_multiple_files=True)

        if uploaded_files:
            # --- NEW PREVIEW SECTION ---
            st.write("📸 **Uploaded Previews:**")
            preview_cols = st.columns(4)
            for idx, uf in enumerate(uploaded_files):
                with preview_cols[idx % 4]:
                    st.image(Image.open(uf), use_container_width=True, caption=f"File {idx+1}")
            
            st.divider()
            
            if st.button("🔍 Extract & Process All", type="primary", use_container_width=True):
                # Phase 1: Scan files
                st.subheader("📋 Scanning Files...")
                status_container = st.container()
                extracted_items = []
                
                for i, uf in enumerate(uploaded_files, 1):
                    with st.spinner(f"Processing {uf.name}..."):
                        result = extract_from_image(uf)
                    
                    col_icon, col_txt = st.columns([1, 4])
                    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    if result:
                        result["filename"] = uf.name
                        extracted_items.append(result)
                        with col_icon:
                            st.image(Image.open(uf).resize((60, 60)), use_container_width=False)
                        with col_txt:
                            st.write(f"✅ **{uf.name}**\n`{result['product_code']}`")
                    else:
                        # LOG FAILURE TO HISTORY
                        with col_icon:
                            st.error("❌")
                        with col_txt:
                            st.write(f"**{uf.name}**\n*Failed to extract*")
                        
                        log_data = pd.DataFrame([{
                            "Timestamp": current_time,
                            "User": st.session_state["user"],
                            "Type": transaction_type,
                            "Product Code": "N/A",
                            "Product Name": "Extraction Failed",
                            "Category": "N/A",
                            "Qty Diff": 0,
                            "Status": "Failed",
                            "Reason": "OCR/LLM Failed"
                        }])
                        if os.path.exists(TXN_FILE):
                            log_data.to_csv(TXN_FILE, mode='a', header=False, index=False)
                        else:
                            log_data.to_csv(TXN_FILE, index=False)

                if extracted_items:
                    # Phase 2: Process table
                    st.divider()
                    st.subheader(f"📦 Processed Product Codes ({len(extracted_items)}/{len(uploaded_files)})")
                    display_data = []
                    for item in extracted_items:
                        qty_change = item["quantity"] if transaction_type == "Incoming" else -item["quantity"]
                        display_data.append({
                            "File": item["filename"], "Product Code": item["product_code"], "Product Name": item["product_name"],
                            "Category": item["category"], "Qty Change": f"{'+' if qty_change > 0 else ''}{qty_change}",
                            "Base Price": f"₹{item['base_price']}", "MRP": item["mrp"] if item["mrp"] else "N/A", "Package": item["package"]
                        })
                    st.dataframe(pd.DataFrame(display_data), use_container_width=True, hide_index=True)

                    # Save to DB
                    with st.spinner("Recording & updating inventory..."):
                        all_sheets = {}
                        inv_cols = ["Product Code", "Product Name", "Quantity", "Base Price", "MRP", "Package"]
                        if os.path.exists(EXCEL_FILE):
                            try:
                                xls_inv = pd.ExcelFile(EXCEL_FILE)
                                for sn in xls_inv.sheet_names:
                                    if sn.lower() not in ("sheet1", "sheet 1"):
                                        all_sheets[sn] = pd.read_excel(xls_inv, sheet_name=sn)
                            except Exception: pass

                        errors = []
                        for item in extracted_items:
                            p_code = item["product_code"]
                            p_name = item["product_name"]
                            qty_change = item["quantity"] if transaction_type == "Incoming" else -item["quantity"]
                            sheet_name = item["category"] if item["category"] != "N/A" else "Uncategorized"
                            status = "Success"
                            reason = ""

                            if item["category"] == "N/A":
                                status = "Partial"
                                reason = "Not in Price List"

                            if sheet_name not in all_sheets:
                                all_sheets[sheet_name] = pd.DataFrame(columns=inv_cols)
                            df = all_sheets[sheet_name]

                            if str(p_code) in df["Product Code"].astype(str).values:
                                idx = df.index[df["Product Code"].astype(str) == str(p_code)].tolist()[0]
                                new_inv_qty = df.at[idx, "Quantity"] + qty_change
                                if new_inv_qty < 0:
                                    errors.append(f"'{p_code}': Out of stock")
                                    status = "Failed"
                                    reason = "Insufficient Stock"
                                else:
                                    df.at[idx, "Quantity"] = new_inv_qty
                                    df.at[idx, "Product Name"] = p_name
                            else:
                                if qty_change < 0:
                                    errors.append(f"'{p_code}': Not in inventory")
                                    status = "Failed"
                                    reason = "Not in Stock"
                                else:
                                    df = pd.concat([df, pd.DataFrame([{"Product Code": p_code, "Product Name": p_name, "Quantity": qty_change, "Base Price": item["base_price"], "MRP": item["mrp"], "Package": item["package"]}])], ignore_index=True)

                            all_sheets[sheet_name] = df
                            
                            # Log Transaction
                            log_data = pd.DataFrame([{
                                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "User": st.session_state["user"],
                                "Type": transaction_type,
                                "Product Code": p_code, "Product Name": p_name, "Category": sheet_name,
                                "Qty Diff": qty_change, "Status": status, "Reason": reason
                            }])
                            if os.path.exists(TXN_FILE):
                                log_data.to_csv(TXN_FILE, mode='a', header=False, index=False)
                            else:
                                log_data.to_csv(TXN_FILE, index=False)

                        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                            for sn, sdf in all_sheets.items():
                                sdf.to_excel(writer, sheet_name=sn[:31], index=False)

                    if errors:
                        for e in errors: st.error(f"⚠️ Skipped {e}")
                    st.balloons()
                    st.success(f"✅ Finished Processing!")
# 6. MANAGE MASTER LIST (Admin Only)
elif current_page == "Manage Master List":
    st.title("🛠️ Manage Master Price List")
    st.info("Edit your 'Golden Database' here. Changes are saved back to the master Excel file.")
    
    xls_path = "priceListHomeFurniture.xlsx"
    if not os.path.exists(xls_path):
        st.error("Master Excel file not found!")
        st.stop()
        
    xls = pd.ExcelFile(xls_path)
    sheet_name = st.selectbox("Select Sheet to Edit", xls.sheet_names)
    
    # Read with header=5 to preserve the 6-row offset
    df = pd.read_excel(xls, sheet_name=sheet_name, header=5)
    
    st.write(f"📝 **Editing: {sheet_name}**")
    edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True, key=f"editor_{sheet_name}")
    
    if st.button("💾 Save Changes to Excel", type="primary", use_container_width=True):
        try:
            with st.spinner("Updating master database..."):
                from openpyxl import load_workbook
                
                # 1. Save data to a temporary in-memory buffer
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    edited_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 2. Open original workbook and replace data from Row 7 (index 6, header=5) down
                wb = load_workbook(xls_path)
                ws = wb[sheet_name]
                
                # Clear rows from 7 downwards (preserving top 6 rows)
                for row in ws.iter_rows(min_row=7):
                    for cell in row: cell.value = None
                
                # Write Header and Data
                # Headers at Row 6 (Excel indexing is 1-based, so Row 6)
                for c_idx, col in enumerate(edited_df.columns, 1):
                    ws.cell(row=6, column=c_idx, value=col)
                
                # Data starting at Row 7
                for r_idx, row in enumerate(edited_df.values, 7):
                    for c_idx, val in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=val)
                        
                wb.save(xls_path)
                
                # 3. Clear Cache to force reload
                if os.path.exists(PRICE_CACHE): os.remove(PRICE_CACHE)
                st.cache_data.clear()
                
                st.success(f"Successfully updated '{sheet_name}' in master Excel file!")
                st.balloons()
        except Exception as e:
            st.error(f"Save Failed: {e}")
